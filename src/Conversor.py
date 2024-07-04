from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
import csv

class Conversor:    
    def convertir_csv_xlsx(rutaArchivo, rutaConvertido):
        wb = Workbook()
        ws = wb.active
        anchoCol = 14.5

        with open(rutaArchivo, 'rt') as fp:
            for row in csv.reader(fp, delimiter=';'):
                ws.append(row)
    
        for rows in ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=1):
           
            # Comprueba los distintos escenarios para saber de qué color debe ser la fila.
            if(rows[0].coordinate == 'A1'):
                color = "C0C0C0"
            elif((rows[0].row == ws.max_row)):
                color = 'FFFF00'
            elif(rows[0].coordinate == 'A2' or (rows[0].value != '' and ws[rows[0].column_letter + str(rows[0].row - 1)].value == '')):
                color = 'FFCC00'
            elif((rows[0].value == '') and ws[rows[0].column_letter + str(rows[0].row - 1)].value != ''):
                color = 'CCCCFF'
            else:
                color = ''
            
            # En el caso de que se haya asignado un color, pinta toda la fila.
            if(color != ''):
                for cell in rows:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))

        for col in ws.iter_cols():
            ws.column_dimensions[col[0].column_letter].width = anchoCol

        wb.save(rutaConvertido)
