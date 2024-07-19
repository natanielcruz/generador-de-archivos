import openpyxl
import io

# Genera un archivo .csv para cada cliente que tenga más de $30000 de deuda y 10 días de atraso. 
# Recibe la direccion del archivo Excel y la ruta donde se guardarán los archivos, además de los colores para identificar los clientes.

class GeneradorDeArchivos:
    def __init__(self, rutaWb, rutaArchivos, colorInicial, colorFinal):
        self.rutaWb = rutaWb
        self.rutaArchivos = rutaArchivos
        self.colorIncial = colorInicial
        self.colorFinal = colorFinal

    def generarArchivos(self):
        with open(self.rutaWb, "rb") as f:
            in_mem_file = io.BytesIO(f.read())

        wb = openpyxl.load_workbook(in_mem_file, data_only=True, read_only=False)
        ws = wb.active
        COLUM_MAX = 11
        header = ('Cliente;Descripcion;Comprobante;N.Recibo/Factura;Dias;F.Base;F.Vencimiento;Imp.Aplicado;Imp.Documento;Atraso;Interes') # Define el encabezado de las tablas del Excel a procesar.
        celdaIni = 0


        # En estos for anidados, se iteran las filas y sus celdas. Cuando encuentra una celda del colorInicial, cambia el valor de la bandera celdaInicial.
        # Cuando llega a la última fila (identificada por colorFinal) y la columna K (que es el monto que nos importa conocer) revisa que se cumplan las condiciones.
        # Si el importe es mayor a $30000 y cuenta con más de 10 días de atraso, se crea un archivo .csv donde se guardará la tabla del cliente.
        for fila in ws.iter_rows(max_col = COLUM_MAX + 1):
            fp = 0
            for celda in fila:
                if(celda.fill.fgColor.rgb == self.colorIncial and celdaIni == 0):
                    celdaIni = celda
 
                if(celda.column_letter == 'K' and celda.fill.fgColor.rgb == self.colorFinal):
                    if(celda.value > 30000 and ws['J' + str(celda.row)].value > 10):
                      i = celdaIni.row
                      fp = open(self.rutaArchivos + '\\' + celdaIni.value + ".csv", "wt")
                      fp.write(header + '\n')

                      # En este ciclo se eliminan los caracteres sucios que pueden aparecer al copiar los datos al archivo .csv
                      while i < celda.row + 1:
                          reemplazar = ['None', '[', ']', '\'', '\"']
                          data = str([ws.cell(row=i, column = j).value for j in range(1, COLUM_MAX + 1)])
 
                          for caracter in reemplazar:
                              data = data.replace(caracter, '')

                          data = data.replace(',', ';')
                          fp.write(data + '\n')

                          i += 1

                    celdaIni = 0
                    if (fp != 0):
                        fp.close()
        wb.close()